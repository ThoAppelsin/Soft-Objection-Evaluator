import re, os, shutil
import click
import tarfile
from pathvalidate import sanitize_filepath
# import coverage
import edit_distance
from pylint.reporters.text import TextReporter
from io import StringIO
import tempfile
import vulture
import pandas as pd
from itertools import chain, repeat
import glob
from pathlib import Path
from contextlib import redirect_stderr
from alive_progress import alive_it
import tokenize
import mergedeep
import warnings
from multiprocessing import Pool, cpu_count
import openpyxl
from copy import copy


def tarextract(tar, outdir):
	with tarfile.open(tar) as tf:
		for file in tf:
			file.name = sanitize_filepath(file.name, replacement_text="_")

		tf.extractall(outdir)


def tarsextract(tars, outdir):
	if outdir.is_dir():
		if click.confirm(f"Directory '{outdir.relative_to(outdir.parents[3])}' already exists, want to delete it and extract new?", default=False):
			shutil.rmtree(outdir)
		else:
			return

	outdir.mkdir(parents=True)

	arguments = list(zip(tars, repeat(outdir)))
	with Pool(min(cpu_count(), len(arguments))) as pool:
		pool.starmap(tarextract, arguments)


def sanitize(code, codepath, full=False):
	resourceindicator = '/'.join(codepath.parts[-6:-2]) + (' (full)' if full else '')

	def join_triplequote_strings(code):
		rline = ''
		inquotes = False
		for line in code:
			rline += line
			qactivity = True
			while qactivity:
				qactivity = False
				if inquotes:
					qend = line.find(inquotes)
					if qend != -1:
						line = line[qend+len(inquotes):]
						inquotes = False
						qactivity = True
				else:
					qstart = sorted((s, q) for q in ["'''", '"""'] if (s := line.find(q)) != -1)
					if qstart:
						s, q = qstart[0]
						line = line[s+len(q):]
						inquotes = q
						qactivity = True
			if inquotes:
				if not rline.endswith('\\'):
					rline += '\\n'
			else:
				yield rline
				rline = ''


	def comment_index(line):
		try:
			comments = [x for x in tokenize.generate_tokens(StringIO(line).readline) if x[0] == tokenize.COMMENT]
			return comments[0][2][1] if len(comments) == 1 else -1
		except tokenize.TokenError as e:
			if line.find('#') != -1:
				print(f"[{resourceindicator}] We have a # and:", str(e))
			return -1


	def join_lines(code):
		rline = ''
		for line in code:
			rline += line
			if comment_index(line) == -1 and rline.endswith('\\'):
				rline = rline[:-1]
			else:
				yield rline
				rline = ''


	# def remove_triplequote_comments(code):
	# 	incomment = False
	# 	commenttype = ''
	# 	for line in code:
	# 		if incomment:
	# 			if line.find(commenttype * 3) >= 0:
	# 				incomment = False
	# 		else:
	# 			l = line
	# 			found_inline_comment = True
	# 			while found_inline_comment:
	# 				found_inline_comment = False
	# 				l = l.lstrip()
	# 				for ctype in ['"', "'"]:
	# 					if l.startswith(ctype * 3):
	# 						cend = l[3:].find(ctype * 3)
	# 						if cend == -1:
	# 							incomment = True
	# 							commenttype = ctype
	# 							break
	# 						else:
	# 							l = l[cend+6:]
	# 							found_inline_comment = True
	# 							break
	# 			if not incomment and l.strip() != '':
	# 				yield line


	def remove_quote_comments(code):
		for line in code:
			l = line.strip()
			repeat = True
			while repeat:
				repeat = False
				for q in ["'''", '"""', "'", '"']:
					if l.startswith(q):
						qend = l.find(q, len(q))
						if qend == -1:
							print(f"[{resourceindicator}] Quotes left unclosed! Full line: >{line}< and rest of the line >{l}")
							break
						l = l[qend+len(q):].lstrip()
						repeat = True
						break
			if l != '':
				yield line


	def remove_comments(code):
		return (line if (ci := comment_index(line)) == -1 else line[:ci] for line in code)


	def rstrip(code):
		return (line.rstrip(" \t\n\r;") for line in code)


	def remove_empty_lines(code):
		return (line for line in code if line != '')

	return list(remove_empty_lines(rstrip(remove_comments(remove_quote_comments(join_lines(join_triplequote_strings(code)))))))


def get_comment(line):
	start = line.find('#')
	return "" if start == -1 else line[start+1:]


def get_comments(code):
	return (get_comment(line) for line in code)


def extract_user_code(code):
	bflag = 'DO_NOT_EDIT_ANYTHING_ABOVE_THIS_LINE'
	eflag = 'DO_NOT_EDIT_ANYTHING_BELOW_THIS_LINE'

	flags = (True if bflag in c else False if eflag in c else None for c in get_comments(code))
	flags = [(i, f) for i, f in enumerate(flags) if f != None]
	goodflags = all(x[1] for x in flags[::2]) and not any(x[1] for x in flags[1::2]) and len(flags) % 2 == 0

	def sturanges():
		rangestart = -1
		for i, f in flags:
			if rangestart == -1:
				if f:
					rangestart = i + 1
			else:
				if not f:
					yield (rangestart, i)
					rangestart = -1

	return [line for r in sturanges() for line in code[r[0]:r[1]]], goodflags


def calculate_edit_distance(old, new):
	sm = edit_distance.SequenceMatcher(old, new)
	return sm.distance()


def run_pylint(testfpath):
	from pylint import lint

	PERFECT = "\n------------------------------------\nYour code has been rated at 10.00/10\n\n"
	ARGS = ["--disable=all", "--enable=W0104,W0105"]
	outIO = StringIO()
	lint.Run([testfpath]+ARGS, reporter=TextReporter(outIO), exit=False)

	outIO.seek(0)
	output = outIO.read()

	return output == PERFECT or output == "" or output


def prepare_vulture_whitelist(srcpath):
	wlpath = srcpath.parent / (srcpath.stem + '-whitelist.py')

	v = vulture.Vulture()
	v.scavenge([srcpath])
	with open(wlpath, 'w') as wlf:
		print('\n'.join(item.get_whitelist_string() for item in v.get_unused_code()), file=wlf)

	return wlpath


def run_vulture(testfpath, vulturewlpath):
	v = vulture.Vulture()
	with redirect_stderr(StringIO()) as f:
		v.scavenge([testfpath, vulturewlpath])
		errout = f.getvalue()
	vulpath = vulture.utils.format_path(Path(testfpath).resolve())
	invalid_syntax_regex = str(vulpath).replace('\\', '\\\\') + r":\d+: invalid syntax at .*"
	errout = '\n'.join(l for l in errout.splitlines() if not re.match(invalid_syntax_regex, l))
	output = [errout] if errout else []
	output += [item.get_report() for item in v.get_unused_code()]
	return output == [] or '\n'.join(output)


def run_tests(code, vulturewlpath, prepend=''):
	temp = tempfile.NamedTemporaryFile(mode="w", delete=False)
	temp.write('\n'.join(code))
	testfpath = temp.name
	temp.close()

	pylintres = run_pylint(testfpath)
	vultureres = run_vulture(testfpath, vulturewlpath)

	os.remove(testfpath)

	if prepend:
		prepend += '-'
	return {prepend + 'pylint': pylintres, prepend + 'vultur': vultureres}


def num_colon_follow(code):
	return sum(0 <= line.find(':') < len(line)-1 for line in code)


def num_semicolon(code):
	return '\n'.join(code).count(';')


def num_comma(code):
	def num_comma_in_line(line):
		try:
			count = 0
			comma_counting_at_level = [True]
			may_start_function = False

			for token in tokenize.generate_tokens(StringIO(line).readline):
				if token.type == tokenize.OP:
					if token.string == ',' and comma_counting_at_level[-1]:
						count += 1
					elif token.string == '(':
						comma_counting_at_level.append(not may_start_function)
					elif token.string == ')':
						comma_counting_at_level.pop()
				may_start_function = token.type == tokenize.NAME
			return count
		except tokenize.TokenError as e:
			return line.count(',') # len(re.findall(r""",\s*(?!"|'|\s|end\s*=)""", line))

	return sum(num_comma_in_line(line) for line in code)


def num_exec(code):
	return '\n'.join(code).count('exec(')


def num_global_nonlocal(code):
	return '\n'.join(code).count('global') + '\n'.join(code).count('nonlocal')


def num_ternary(code):
	return sum(len(re.findall(r"""\bif\b.*\belse\b""", line)) for line in code)


def num_multi_assign(code):
	return sum(max(0, len(re.findall(r"""[^=!+\-*\^\|&<>%/]=[^=]""", line)) - 1) for line in code)


def num_self_assign(code):
	return sum(1 if re.match(r"""^\s*(\w+)\s*=\s*\1\s*$""", line) else 0 for line in code)


def num_empty_string_return(code):
	return sum(1 if re.match(r"""^\s*return\s*("\s*"|'\s*'|\("\s*"\)|\('\s*'\))\s*$""", line) else 0 for line in code)


def num_silly_and_or(code):
	aoexp = r"""\b(and|or)\b"""
	strexp = r"""\s*("[^"]*"|'[^']*')\s*"""
	rendexp = r"""(:|\)|\bor\b|\band\b)"""
	lendexp = r"""(if\b|\(|\bor\b|\band\b|^\s*)"""
	return sum(len(re.findall(aoexp + strexp + rendexp + "|" + lendexp + strexp + aoexp, line)) for line in code)


def num_stray_and_or(code):
	return sum(len(re.findall(r"""^(?:(?!if)(?!while).)*?\b(and|or)\b.*""", line)) for line in code)


# class MultiAssignCountVisitor(ast.NodeVisitor):
# 	def __init__(self, *args, **kwargs):
# 		super().__init__(*args, **kwargs)
# 		self.count = 0

# 	def visit_Assign(self, node):
# 		if len(node.targets) > 1 or any(type(t) is ast.Tuple for t in node.targets):
# 			self.count += 1


# def num_multi_assign(code):
# 	macv = MultiAssignCountVisitor()
# 	ap = ast.parse('\n'.join(code))
# 	macv.visit(ap)
# 	return macv.count


# def num_multi_assign_naive(code):
# 	return sum(0 if (eqidx := line.find('=')) == -1 else line[:eqidx].count(',') for line in code)


flawless = {
	'org-#colfol': 0,
	'org-#semcol': 0,
	'org-#comma': 0,
	'org-#exec': 0,
	'org-#mulas': 0,
	'org-#globl': 0,
	'org-#tern': 0,
	'org-#selas': 0,
	'org-#esret': 0,
	'org-#silao': 0,
	'cor-#sryao': 0,
	'org-flagOK': True,
	'org-pylint': True,
	'org-vultur': True,
	'cor-#colfol': 0,
	'cor-#semcol': 0,
	'cor-#comma': 0,
	'cor-#exec': 0,
	'cor-#mulas': 0,
	'cor-#globl': 0,
	'cor-#tern': 0,
	'cor-#selas': 0,
	'cor-#esret': 0,
	'cor-#silao': 0,
	'cor-#sryao': 0,
	'cor-flagOK': True,
	'cor-pylint': True,
	'cor-vultur': True
	}


def get_flaws(report):
	return ', '.join(k for k in flawless if k in report and flawless[k] != report[k])


def get_report(orgpath, corpath, vulturewlpath, should_sanitize=True):
	with open(orgpath) as orgfile:
		orgfull = orgfile.read().splitlines()
		org, orggoodflags = extract_user_code(orgfull)
	with open(corpath) as corfile:
		corfull = corfile.read().splitlines()
		cor, corgoodflags = extract_user_code(corfull)
	if should_sanitize:
		orgfull = sanitize(orgfull, orgpath, True)
		corfull = sanitize(corfull, corpath, True)
		org = sanitize(org, orgpath)
		cor = sanitize(cor, corpath)

	if len(cor) > 0:
		orgreport = {
			'org-#lines' : len(org),
			'org-#colfol': num_colon_follow(org),
			'org-#semcol': num_semicolon(org),
			'org-#comma': num_comma(org),
			'org-#exec': num_exec(org),
			'org-#mulas': num_multi_assign(org),
			'org-#globl': num_global_nonlocal(org),
			'org-#tern': num_ternary(org),
			'org-#selas': num_self_assign(org),
			'org-#esret': num_empty_string_return(org),
			'org-#silao': num_silly_and_or(org),
			'org-#sryao': num_stray_and_or(org),
			'org-flagOK': orggoodflags,
			} | run_tests(orgfull, vulturewlpath, 'org')
		correport = {
			'cor-#lines': len(cor),
			'cor-#colfol': num_colon_follow(cor),
			'cor-#semcol': num_semicolon(cor),
			'cor-#comma': num_comma(cor),
			'cor-#exec': num_exec(cor),
			'cor-#mulas': num_multi_assign(cor),
			'cor-#globl': num_global_nonlocal(cor),
			'cor-#tern': num_ternary(cor),
			'cor-#selas': num_self_assign(cor),
			'cor-#esret': num_empty_string_return(cor),
			'cor-#silao': num_silly_and_or(cor),
			'cor-#sryao': num_stray_and_or(cor),
			'cor-flagOK': corgoodflags
			} | run_tests(corfull, vulturewlpath, 'cor')
		report = {'edit_dist': calculate_edit_distance(org, cor)} | orgreport | correport
		return report, get_flaws(orgreport) == "", get_flaws(correport) == ""
	else:
		return False


def subreport(report, tag):
	return {k: v for k, v in report.items() if (k.startswith(f"{tag}-") if tag else "-" not in k)}


def collect_gradebook(path, suffix):
	with warnings.catch_warnings(record=True):
		gradebook = pd.read_excel(path, header=1)

	gradecolumns = [c for c in gradebook.columns if c.startswith('Total')]

	if len(qlists := gradebook["Question Id List"].unique()) != 1:
		print(f"Gradebook {path.name} contains multiple Question Id Lists: {qlists}")

	return ({'user': f"user{r['User ID']}", 'qid': f"question{q}", f'grade-{suffix}': g, f'gbook-{suffix}': f'=HYPERLINK("{path}")'}
			for i, r in gradebook.iterrows()
			for q, g in zip(str(r["Question Id List"]).split(", "), r[gradecolumns]))


def create_nppath_backref(ppath, nppath):
	brefpath = ppath.parent / 'backref.txt'
	with open(brefpath, "w") as bref:
		bref.write(str(nppath))


def consider_creating_patch(isperfect, ppath, path):
	if isperfect or (ppath.exists() and path.samefile(ppath)):
		return path

	ppath.parent.mkdir(parents=True)
	shutil.copyfile(path, ppath)
	create_nppath_backref(ppath, path)
	return ppath


def analyze_stuq(examid, stuid, oqid, npcpath, pcpath, npopath, popath, reportworthy, vulturewlpath):
	if reportworthy:
		opath = popath if popath.is_file() else npopath
		cpath = pcpath if pcpath.is_file() else npcpath
		reportpack = get_report(opath, cpath, vulturewlpath)
		if reportpack:
			report, orgtestperfect, cortestperfect = reportpack
			opath = consider_creating_patch(orgtestperfect, popath, opath)
			cpath = consider_creating_patch(cortestperfect, pcpath, cpath)
			return {
				'user': stuid,
				# 'qnum': ns.origqiddict[oqid]['qnum'],
				'qid': oqid,
				# 'sect': ns.correctiondict[stuid][oqid]['section'],
				# 'exam': examid,
				'org': f'=HYPERLINK("{opath}")',
				'org*': f'=HYPERLINK("{npopath}")' if opath == popath else None
				} | subreport(report, 'org') | {
				'cor': f'=HYPERLINK("{cpath}")',
				'cor*': f'=HYPERLINK("{npcpath}")' if cpath == pcpath else None
				} | subreport(report, 'cor') | subreport(report, False)

	return {
		'user': stuid,
		'qid': oqid,
		'org': f'=HYPERLINK("{opath}")',
		'cor': f'=HYPERLINK("{cpath}")'
		}


def format_excel(path, freezerows, freezecolumns):
	wb = openpyxl.load_workbook(filename=path)
	ws = wb.active
	ws.freeze_panes = f"{openpyxl.utils.get_column_letter(freezecolumns + 1)}{freezerows + 1}"
	for row in ws.iter_rows(1, ws.max_row, 1, ws.max_column):
		for cell in row:
			alignmentstyle = copy(cell.alignment)
			alignmentstyle.shrinkToFit = True
			cell.alignment = alignmentstyle
	ws.auto_filter.ref = f"A{freezerows}:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}"
	wb.save(path)



if __name__ == '__main__':
	CURRENT_EXAM = 1

	coursehome = Path.home() / "Downloads/cmpe150fall2022"

	if CURRENT_EXAM == 1:
		examname = "mt1"
		have_legitrange = False
		origqiddict = {
			'question1291': {'qnum': 'q1', 'corrid': 'question1291'},
			'question1292': {'qnum': 'q1', 'corrid': 'question1292'},
			'question1293': {'qnum': 'q2', 'corrid': 'question1293'},
			'question1294': {'qnum': 'q2', 'corrid': 'question1294'},
			'question1295': {'qnum': 'q3', 'corrid': 'question1295'},
			'question1296': {'qnum': 'q3', 'corrid': 'question1296'}
			}
	elif CURRENT_EXAM == 2:
		examname = "mt2"
		have_legitrange = True
		origqiddict = {
			'question1327': {'qnum': 'q1', 'corrid': 'question1336', 'legitrange': (6, 15)},
			'question1328': {'qnum': 'q2', 'corrid': 'question1337', 'legitrange': (11, 25)},
			'question1329': {'qnum': 'q3', 'corrid': 'question1338', 'legitrange': (8, 20)},
			'question1330': {'qnum': 'q1', 'corrid': 'question1339', 'legitrange': (6, 15)},
			'question1331': {'qnum': 'q2', 'corrid': 'question1340', 'legitrange': (11, 25)},
			'question1332': {'qnum': 'q3', 'corrid': 'question1341', 'legitrange': (8, 20)}
			}

	corrqiddict = {v['corrid'] : {'qnum': v['qnum'], 'origid': oqid} for oqid, v in origqiddict.items()}

	examhome = coursehome / examname

	rawhome = examhome / "raw"
	rawquestionshome = rawhome / "questions"
	raworiginalshome = rawhome / "originals"
	rawcorrectionshome = rawhome / "corrections"

	rawquestiontars = rawquestionshome.glob("*.tar.gz")
	raworiginaltars = raworiginalshome.glob("*.tar.gz")
	rawcorrectiontars = rawcorrectionshome.glob("*/*.tar.gz")

	processedhome = examhome / "processed"
	processedquestionsdir = processedhome / "questions"
	processedoriginalsdir = processedhome / "originals"
	processedcorrectionsdir = processedhome / "corrections"

	patchhome = examhome / "patch"
	patchoriginalsdir = patchhome / "originals"
	patchcorrectionsdir = patchhome / "corrections"


	# EXTRACT TARS
	tarsextract(rawquestiontars, processedquestionsdir)
	tarsextract(raworiginaltars, processedoriginalsdir)
	tarsextract(rawcorrectiontars, processedcorrectionsdir)

	# PREPARE VULTURE WHITELISTS
	vulturewldict = {qdir.name : prepare_vulture_whitelist(qdir / 'src/Main.py') for qdir in processedquestionsdir.iterdir()}

	# PREPARE POINTERS TO CORRECTIONS
	correctiondict = mergedeep.merge({}, *({ cpath.parts[-4] : { cpath.parts[-3] : { 'path': cpath, 'section': cpath.parts[-5].split('_')[1] } } }
											for cpath in processedcorrectionsdir.glob("*/*/*/src/Main.py")))

	# COLLECT ORIGINAL GRADES
	originalgradebooks = raworiginalshome.glob("*.xlsx")
	originaldf = pd.DataFrame(chain(*(collect_gradebook(gb, 'org') for gb in originalgradebooks))).set_index(['user', 'qid']).sort_index()

	# COLLECT CORRECTION GRADES
	correctiongradebooks = rawcorrectionshome.glob("*/*.xlsx")
	correctiondf = pd.DataFrame(chain(*(collect_gradebook(gb, 'cor') for gb in correctiongradebooks)))
	correctiondf['qid'] = correctiondf['qid'].apply(lambda x: corrqiddict[x]['origid'])
	correctiondf = correctiondf.set_index(['user', 'qid']).sort_index()
	for pcgpath in patchcorrectionsdir.glob("*/*/src/grade.txt"):
		with open(pcgpath) as pcgf:
			correctiondf.loc[(pcgpath.parts[-4], corrqiddict[pcgpath.parts[-3]]['origid']), 'grade-cor'] = int(pcgf.readline())
	correctiondf = correctiondf.sort_values('grade-cor', ascending=False).groupby(['user', 'qid']).first().sort_index()

	# COLLECT STUDENT INFO
	studentinfodf = pd.read_excel(coursehome / 'studentinfo.xlsx')
	studentinfodf['user'] = ['user' + str(x) for x in studentinfodf['user']]
	studentinfodf = studentinfodf.set_index('user')
	studentinfodf.columns = pd.MultiIndex.from_product([['INFO'], studentinfodf.columns])

	# PREPARE REPORT

	opaths = list(processedoriginalsdir.glob("*/*/*/src/Main.py"))


	def patchpath(patchdir, stuid, qid):
		return patchdir / stuid / qid / "src/Main.py"


	def handle_patches(patchdir, stuid, qid, nppath):
		ppath = patchpath(patchdir, stuid, qid)
		if ppath.is_file():
			create_nppath_backref(ppath, nppath)
		
		return ppath


	def produce_arguments():
		for npopath in (bar := alive_it(opaths)):
			examid = npopath.parts[-5].split('_')[1]
			stuid = npopath.parts[-4]
			oqid = npopath.parts[-3]
			cqid = origqiddict[oqid]['corrid']

			bar.title(f'on {stuid}-{oqid}')

			if cqid in correctiondict[stuid]:
				npcpath = correctiondict[stuid][cqid]['path']
				pcpath = handle_patches(patchcorrectionsdir, stuid, cqid, npcpath)
				popath = handle_patches(patchoriginalsdir, stuid, oqid, npopath)

				reportworthy = True # (stuid, oqid) in correctiondf.index and correctiondf.loc[(stuid, oqid), 'grade-cor'].item() > 0

				yield examid, stuid, oqid, npcpath, pcpath, npopath, popath, reportworthy, vulturewldict[oqid]


	with Pool() as pool:
		# arguments = list(produce_arguments())
		# results = pool.starmap(analyze_stuq, tqdm(arguments))
		results = pool.starmap(analyze_stuq, produce_arguments())
		reportdf = pd.DataFrame(results)

	if have_legitrange:
		reportdf['ratio'] = reportdf.apply(lambda r: 1 if r['edit_dist'] == 0 else max(0, min(
			1 - r['edit_dist'] / min(max(r['org-#lines'], r['cor-#lines']), origqiddict[r['qid']]['legitrange'][1]),
			(max(r['org-#lines'], r['cor-#lines']) - r['edit_dist']) / max(r['org-#lines'], r['cor-#lines'], origqiddict[r['qid']]['legitrange'][0])
			)), axis=1)
		# reportdf['ratio'] = np.where(reportdf['edit_dist'] == 0, 1, pd.concat((1 - reportdf['edit_dist'] / reportdf[['org-#lines', 'cor-#lines']].assign(legitmax=.max(axis=1))
	else:
		reportdf['ratio'] = 1 - reportdf['edit_dist'] / reportdf[['org-#lines', 'cor-#lines']].max(axis=1)

	for pf in ('org', 'cor'):
		reportdf[f'{pf}-inspect'] = reportdf[[c for c in reportdf.columns if c.startswith(pf)]].apply(get_flaws, axis=1)
	reportdf['all-inspect'] = reportdf.apply(get_flaws, axis=1)

	reportcorrectionspath = examhome / f'report_corrections_{examname}.xlsx'
	reportdf[( c for c in reportdf.columns if c not in ['qid', 'sect', 'exam'] and (c not in flawless or (reportdf[c] != flawless[c]).any()) )].to_excel(reportcorrectionspath)
	format_excel(reportcorrectionspath, 1, 1)

	# reportdf.pivot(index="user", columns="qnum").swaplevel(0, 1, axis=1).sort_index(1)['q2']

	df = originaldf.join(correctiondf).join(reportdf.set_index(['user', 'qid'])).reset_index(1)
	df['qnum'] = df['qid'].apply(lambda x: origqiddict[x]['qnum'])
	df['grade-new'] = pd.concat([df['grade-org'], df['grade-cor'] * df['ratio']], axis=1).max(axis=1)

	qnums = df['qnum'].unique()
	df = df.pivot(columns='qnum').swaplevel(0, 1, axis=1).sort_index(axis=1)
	df[('TOTAL', 'ORIGINAL')] = pd.concat((df[(qnum, 'grade-org')] for qnum in qnums), axis=1).mean(axis=1)
	df[('TOTAL', 'NEW')] = pd.concat((df[(qnum, 'grade-new')] for qnum in qnums), axis=1).mean(axis=1)
	df[('TOTAL', 'DELTA')] = df[('TOTAL', 'NEW')] - df[('TOTAL', 'ORIGINAL')]
	# df[('INFO', 'STUDENT ID')] = [studentinfodf.loc[user, 'studeintID'] if user in studentinfodf.index else 'MISSING' for user in df.index]
	df = df.join(studentinfodf)

	reportfullpath = examhome / f'report_full_{examname}.xlsx'
	df.to_excel(reportfullpath)

	format_excel(reportfullpath, 3, 1)
